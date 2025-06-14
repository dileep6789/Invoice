import streamlit as st
from openpyxl import load_workbook
from num2words import num2words
import numpy as np
import os


st.title("Title")
st.header("Input Details")

# Excel file in Downloads folder
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
excel_file = "new_format.xlsx"
sheet_name = "Details"



def number_to_text_indian(number):
    return num2words(number, lang='en_IN')






data = {
     'Beloorbayir Biotech Ltd.,(BC)': {'GST': '29AACCB6949C2ZK', 'Address': '#10,13th Cross,Doddanna Industrial Estate,Adjacent to Peenya 2nd Stage,Bangalore,Karnatka-560091','State':'Karnataka','Code':29},
     'TANUR CORPORATION': {'GST': '24AAEFT4650P1ZS', 'Address': 'Plot No.1 to 50,Main road, Veraval, Gujarat 362266','State':'Gujarat','Code':24},
     'C': {'GST': 'GST789', 'Address': 'Chennai','State':'Karnataka','Code':49}
 }
products ={
     'Chitin':{'HSN': 3913}
 }
Invoice_No=st.text_input("Enter a Invoice No:")
Invoice_Date=st.date_input("Enter a date")
Vechicle_Number=st.text_input("Enter a vechicle")

st.write("Bill to Party")

BillerName=st.selectbox('Select the name',("Beloorbayir Biotech Ltd.,(BC)","TANUR CORPORATION","C","Other"),key='biller_name_selectbox')
if BillerName in data:
    BillerGST = data[BillerName]['GST']
    BillerAddress = data[BillerName]['Address']
    BillerState = data[BillerName]['State']
    BillerCode = data[BillerName]['Code']
    
    st.write(f"**GST:** {BillerGST}")
    st.write(f"**Place of Business:** {BillerAddress}")
    st.write(f"**GST:** {BillerState}")
    st.write(f"**Code:** {BillerCode}")
    
else:
    BillerName=st.text_input("Enter the Name")
    BillerGST = st.text_input("Enter GST")
    BillerAddress = st.text_input("Enter Place of Business")
    BillerState=st.text-input("Enter a state:")
    BillerCode =st.text-input("Enter a code:")
    
st.write("Ship to Party")

ShipperName=st.selectbox('Select the name',("Beloorbayir Biotech Ltd.,(BC)","TANUR CORPORATION","C","Other"),key='shipper_name_selectbox')
if ShipperName in data:
    ShipperGST = data[ShipperName]['GST']
    ShipperAddress = data[ShipperName]['Address']
    ShipperState = data[ShipperName]['State']
    ShipperCode = data[ShipperName]['Code']
    
    st.write(f"**GST:** {ShipperGST}")
    st.write(f"**Place of Business:** {ShipperAddress}")
    st.write(f"**GST:** {ShipperState}")
    st.write(f"**Code:** {ShipperCode}")
    
else:
    ShipperName = st.text_input("Enter the Name")
    ShipperGST = st.text_input("Enter GST")
    ShipperAddress = st.text_input("Enter Place of Business")
    ShipperState=st.text-input("Enter a state:")
    ShipperCode =st.text-input("Enter a code:")

st.write("Product-Details")
product=st.selectbox('Select the Product',("Chitin","Other"))
if product in products:
    ProductName=products[product]['HSN']
    
else:
    ProductName=st.text_input("Enter a product name:")
    
Qty=st.number_input("Enter a Quantity:", min_value=0, step=1)
Rate=st.number_input("Enter a Rate:", min_value=0, step=1)
Amount=Qty*Rate
Discount=st.number_input("Enter a discount")
Taxable= Amount
IGSTAmount=np.round(Taxable*18/100,0)
Total= Taxable + IGSTAmount


Total_in_words=number_to_text_indian(Total)
Total_in_words=Total_in_words.title()

# Save to specific cells
if st.button("Save to Excel"):
    #Load the Excel file and sheet
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        # Write values to fixed cells (change cell addresses if needed)
        ws['A11'] = f"Invoice No: {Invoice_No}"
        ws['A12'] = f"Invoice Date: {Invoice_Date}"
        ws['I12'] = f"Vehicle number: {Vechicle_Number}"
        ws['A17'] = f"Name  :{BillerName}"
        ws['A18'] = f"Address: {BillerAddress}"
        ws['A20']=f"GSTIN: {BillerGST}"
        ws['A21']=f"State: {BillerState}"
        ws['H21']= BillerCode



        ws['I17'] = f"Name  :{ShipperName}"
        ws['I18'] = f"Address: {ShipperAddress}"
        ws['I20']=f"GSTIN: {ShipperGST}"
        ws['I21']=f"State: {ShipperState}"
        ws['P21']=ShipperCode

        ws['B25']=product
        ws['F25']=Qty
        ws['G25']=Rate
        ws['H25']=Amount
        ws['J25']=Taxable
        ws['M25']=IGSTAmount
        ws['O25']=Total

        ws['F33']=Qty
        ws['H33']=Amount
        ws['J33']=Taxable
        ws['M33']=IGSTAmount
        ws['O33']=Total
        ws['A35']=Total_in_words
        ws['O34']=Taxable
        ws['O35']=IGSTAmount
        ws['O36']=Total


        wb.save(excel_file)

        download_excel = "new_format.xlsx"

        # Show Download Excel button
        if os.path.exists(download_excel):
            with open(download_excel, "rb") as f:
                st.download_button("Download Excel", f, file_name=f"Invoice-{Invoice_No}.xlsx", key="excel")
        else:
            st.error("❌ Excel file not found. Make sure it is saved as 'new_format.xlsx'.")
            st.stop()
    
    
